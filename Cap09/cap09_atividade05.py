#https://googlechromelabs.github.io/chrome-for-testing/
#pip install selenium openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from time import sleep
import os
from os import system, name

def clear_screen():
    system('cls') if name == 'nt' else system('clear')

def main():
    produto = input('Informe o produto: ')

    # Configuração para ocultar a janela do navegador (modo headless)
    options = webdriver.ChromeOptions()
    #options.add_argument("--headless") #Oculta a janela do navegador
    navegador = webdriver.Chrome(options=options)

    url = "https://www.mercadolivre.com.br/"
    navegador.get(url)
    sleep(3)

    # Preenche a barra de pesquisa com o produto informado
    barraPesquisa = navegador.find_element(By.NAME, "as_word")
    barraPesquisa.send_keys(produto)
    barraPesquisa.send_keys(Keys.ENTER)
    sleep(2)
  
    clear_screen() # Função Limpa a tela

    # Coleta os dados dos produtos
    listaProdutos = navegador.find_elements(By.CLASS_NAME, "ui-search-layout__item")
    dadosExtraidos = []
    for item in listaProdutos:
        nome = item.find_element(By.CLASS_NAME, "poly-component__title").text
        preco = item.find_element(By.CLASS_NAME, "andes-money-amount__fraction").text
        preco = preco.replace('.', '')  # Remove os pontos do preço
        urlProduto = item.find_element(By.TAG_NAME, "a").get_attribute("href")
        dadosExtraidos.append((nome, float(preco), urlProduto))

    navegador.quit()
    try:
        # Atualiza o arquivo Excel com os dados extraídos
        arq = load_workbook("pesquisaPreço.xlsx")
        data_atual = datetime.now().strftime('%d-%m-%Y')
        plan = f'{data_atual}-{produto}'
        arq.create_sheet(plan)
        planAtiva = arq[plan]
        for linha in dadosExtraidos:
            planAtiva.append(linha)

        arq.save("pesquisaPreço.xlsx")
        print('Dados extraídos e gravados com sucesso!!!')
    except:
        print('Erro inesperado.')

if __name__ == "__main__":   
    main()
